from tkinter import *
from tkinter import messagebox
import openpyxl



def track():
    category1=cat.get()
    item1=it.get()
    cost1=cos.get()

    cat.delete(0,END)
    it.delete(0, END)
    cos.delete(0, END)

    from openpyxl import load_workbook, Workbook
    import os

    # Function to append data to the Excel file
    def append_to_excel(category1, item1, cost1, filename="spending_tracker.xlsx"):
        # Check if the file exists
        if os.path.exists(filename):
            # Load the existing workbook
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            # Create a new workbook and sheet if the file does not exist
            workbook = Workbook()
            sheet = workbook.active
            # Optionally, add headers
            sheet.append(["Category", "Item", "Cost"])

        # Append the new data
        sheet.append([category1, item1, cost1])

        # Save the workbook
        workbook.save(filename)



    is_ok = True
    if is_ok:
        append_to_excel(category1, item1, cost1)

    if len(category1)==0 or len(item1)==0:
        no_msg = messagebox.showinfo(title="Oops!", message="Please don't leave any field empty!")
    else:
        is_ok = messagebox.askokcancel(title="Info",
                                       message=f"The details entered are\nCategory : {category1}\nItem :"
                                               f" {item1}\nCost : {cost1}")

    if is_ok:
        with open("data.txt", mode="a") as data_file:
            data_file.write(f"{category1}: {item1} {cost1}\n")






window=Tk()
window.title("Spend Tracker")
window.config(padx=40,pady=40)

canvas=Canvas(width=225,height=225)
img=PhotoImage(file="download.png")
canvas.create_image(113,113,image=img)
canvas.grid(row=0,column=1)


button1=Button(text="Track",font=("times",12,"bold"),width=25,command=track)
button1.grid(row=4,column=1,columnspan=2)

category=Label(text="Category",font=("times",18,"bold"),fg="#2980B9")
category.grid(row=1,column=0)

item=Label(text="Item",fg="#2980B9",font=("times",18,"bold"))
item.grid(row=2,column=0)

cost=Label(text="Cost",fg="#2980B9",font=("times",18,"bold"))
cost.grid(row=3,column=0)

cat=Entry(width=25)
cat.focus()
cat.grid(row=1,column=1,columnspan=2)

it=Entry(width=25)
it.grid(row=2,column=1,columnspan=2)

cos=Entry(width=25)
cos.grid(row=3,column=1,columnspan=2)


window.mainloop()
